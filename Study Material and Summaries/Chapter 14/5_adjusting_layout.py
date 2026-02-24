# CHAPTER 14 - SECTION 5: ADJUSTING LAYOUT

# What: Adjust row heights, column widths, merge cells, freeze panes
# Why: Format spreadsheets to be readable and professional
# How: Set dimensions and use merge/freeze methods

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

# Add data
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

# Set row height and column width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

# Merge cells
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'

# Unmerge cells
sheet.unmerge_cells('A1:D3')

# Freeze panes (freeze row 1)
sheet.freeze_panes = 'A2'

wb.save('dimensions.xlsx')
