# CHAPTER 14 - SECTION 2: WRITING EXCEL FILES

# What: Create new Excel files and write data to them
# Why: Generate spreadsheets programmatically instead of by hand
# How: Create Workbook, get sheet, write to cells, save file

import openpyxl

# Creating a new blank workbook
wb = openpyxl.Workbook()
print(wb.sheetnames)  # ['Sheet']

sheet = wb.active
print(sheet.title)  # 'Sheet'

# Changing sheet title
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)  # ['Spam Bacon Eggs Sheet']

# Writing values to cells
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)  # Hello, world!

# Creating and removing sheets
wb.create_sheet()  # Creates 'Sheet1'
print(wb.sheetnames)  # ['Spam Bacon Eggs Sheet', 'Sheet1']

wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)  # ['First Sheet', 'Spam Bacon Eggs Sheet', 'Sheet1']

del wb['Sheet1']  # Delete a sheet
print(wb.sheetnames)  # ['First Sheet', 'Spam Bacon Eggs Sheet']

# Saving the workbook
wb.save('example_output.xlsx')
print('Workbook saved!')
