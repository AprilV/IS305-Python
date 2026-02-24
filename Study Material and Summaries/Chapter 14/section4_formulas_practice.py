# SECTION 4: FORMULAS - Practice

# Q1: Import openpyxl module
# WHAT IT DOES: Makes openpyxl functions available for working with Excel files
# ┌─ EXAMPLE ─────────────
# │ import sys
# └───────────────────────
import openpyxl



# Q2: Use openpyxl.Workbook() to create a new workbook, store workbook in variable wb, and use .active to get the active sheet, store sheet in variable sheet
# WHAT IT DOES: Creates a new empty Excel file in memory and gets its active worksheet
# ┌─ EXAMPLE ─────────────
# │ wb = openpyxl.Workbook()
# │ sheet = wb.active
# └───────────────────────
wb = openpyxl.Workbook()
sheet = wb.active



# Q3: Use sheet['A1'] to write the value 150 to cell A1 and use sheet['A2'] to write 250 to cell A2
# WHAT IT DOES: Writes number values to cells that will be used in formulas
# ┌─ EXAMPLE ─────────────
# │ my_sheet['B1'] = 100
# └───────────────────────
sheet['A1'] = 150
sheet['A2'] = 250



# Q4: Use sheet['A3'] to write the formula '=SUM(A1:A2)' to cell A3
# WHAT IT DOES: Sets a cell to contain an Excel formula that calculates the sum of cells A1 and A2
# ┌─ EXAMPLE ─────────────
# │ my_sheet['B3'] = '=SUM(B1:B2)'
# └───────────────────────
sheet['A3'] =  '=SUM(A1:A2)'



# Q5: Use sheet['B1'] to write value 20 and use sheet['B2'] to write value 4
# WHAT IT DOES: Writes number values to cells in column B for multiplication formula
# ┌─ EXAMPLE ─────────────
# │ my_sheet['C1'] = 15
# │ my_sheet['C2'] = 3
# └───────────────────────
sheet['B1'] = 20
sheet['B2'] = 4



# Q6: Use sheet['B3'] to write the formula '=B1*B2' to cell B3
# WHAT IT DOES: Sets a cell to contain an Excel formula that multiplies B1 and B2
# ┌─ EXAMPLE ─────────────
# │ my_sheet['C3'] = '=C1/C2'
# └───────────────────────
sheet['B3'] = '=B1*B2'


# Q7: Use sheet['C1'] to write value 90 and use sheet['C2'] to write value 85
# WHAT IT DOES: Writes number values to cells in column C for average formula
# ┌─ EXAMPLE ─────────────
# │ my_sheet['D1'] = 100
# └───────────────────────
sheet['C1'] = 90
sheet['C2'] = 85



# Q8: Use sheet['C3'] to write the formula '=AVERAGE(C1:C2)' to cell C3
# WHAT IT DOES: Sets a cell to contain an Excel formula that calculates the average of C1 and C2
# ┌─ EXAMPLE ─────────────
# │ my_sheet['D3'] = '=MAX(D1:D2)'
# └───────────────────────
sheet['C3'] = '=AVERAGE(C1:C2)'


# Q9: Use wb.save() to save the workbook as 'formulas_workbook.xlsx'
# WHAT IT DOES: Writes the workbook from memory to a physical Excel file on disk
# ┌─ EXAMPLE ─────────────
# │ wb.save('calculations.xlsx')
# └───────────────────────
wb.save('formulas_workbook.xlsx')



