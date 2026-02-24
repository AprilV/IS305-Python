# SECTION 5: ADJUSTING LAYOUT - Practice

# Q1: Import openpyxl module
# WHAT IT DOES: Makes openpyxl functions available for working with Excel files
# ┌─ EXAMPLE ─────────────
# │ import sys
# └───────────────────────
import openpyxl



# Q2: Use openpyxl.Workbook() to create a new workbook and use .active to get the active sheet, store sheet in variable sheet
# WHAT IT DOES: Creates a new empty Excel file in memory and gets its active worksheet
# ┌─ EXAMPLE ─────────────
# │ my_wb = openpyxl.Workbook()
# │ my_sheet = my_wb.active
# └───────────────────────
wb = openpyxl.Workbook()
sheet = wb.active



# Q3: Use sheet['A1'] to write 'Product Report', use sheet['B1'] to write 'January', and use sheet['C1'] to write 'February'
# WHAT IT DOES: Writes text values to three adjacent cells in row 1
# ┌─ EXAMPLE ─────────────
# │ my_sheet['A1'] = 'Sales Summary'
# └───────────────────────
sheet['A1'] = 'Product Report'
sheet['B1'] = 'January'
sheet['C1'] = 'February'


# Q4: Use sheet.merge_cells() to merge cells A1 through C1 (use 'A1:C1')
# WHAT IT DOES: Combines multiple cells into one larger cell that spans across columns or rows
# ┌─ EXAMPLE ─────────────
# │ my_sheet.merge_cells('A2:D2')
# └───────────────────────
sheet.merge_cells('A1:C1')



# Q5: Use sheet['A1'] to write 'Monthly Product Sales' (this sets the value for the merged cell)
# WHAT IT DOES: Sets the text content of the merged cell group using the top-left cell coordinate
# ┌─ EXAMPLE ─────────────
# │ my_sheet['A2'] = 'First Quarter Results'
# └───────────────────────
sheet['A1'] = 'Monthly Product Sales'



# Q6: Set sheet.row_dimensions[1].height equal to 35
# WHAT IT DOES: Changes the height of row 1 measured in points (1/72 of an inch)
# ┌─ EXAMPLE ─────────────
# │ my_sheet.row_dimensions[2].height = 25
# └───────────────────────
sheet.row_dimensions[1].height = 35



# Q7: Set sheet.column_dimensions['A'].width equal to 25
# WHAT IT DOES: Changes the width of column A measured in number of characters at default font size
# ┌─ EXAMPLE ─────────────
# │ my_sheet.column_dimensions['B'].width = 15
# └───────────────────────
sheet.column_dimensions['A'].width = 25



# Q8: Set sheet.freeze_panes equal to 'A2' to freeze row 1
# WHAT IT DOES: Locks row 1 so it stays visible when scrolling down through the spreadsheet
# ┌─ EXAMPLE ─────────────
# │ my_sheet.freeze_panes = 'B1'
# └───────────────────────
sheet.freeze_panes = 'A2'



# Q9: Use wb.save() to save the workbook as 'layout_demo.xlsx'
# WHAT IT DOES: Writes the workbook from memory to a physical Excel file on disk
# ┌─ EXAMPLE ─────────────
# │ my_wb.save('formatted_layout.xlsx')
# └───────────────────────
wb.save('layout_demo.xlsx')



