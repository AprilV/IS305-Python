# CHAPTER 14 - SECTION 1: READING EXCEL FILES

# =============================================================================
# WHAT WE'RE LEARNING
# =============================================================================
# How to read data from Excel .xlsx files using the openpyxl module.

# =============================================================================
# WHY THIS MATTERS
# =============================================================================
# Instead of manually opening Excel files and copying data, you can write 
# Python programs that automatically extract information from spreadsheets.
# Perfect for processing reports, analyzing data, or pulling information 
# from hundreds of Excel files.

# =============================================================================
# KEY CONCEPTS
# =============================================================================
# 1. Loading workbooks - Opening Excel files with openpyxl.load_workbook()
# 2. Accessing sheets - Getting specific worksheets by name or the active sheet
# 3. Reading cells - Two ways: sheet['A1'] or sheet.cell(row=1, column=2)
# 4. Cell properties - coordinate, row, column, value
# 5. Sheet dimensions - max_row and max_column tell you how big the data is
# 6. Converting column letters/numbers - get_column_letter() and column_index_from_string()
# 7. Slicing - Getting rectangular areas of cells like sheet['A1':'C3']
# 8. Iterating - Looping through all rows or columns with sheet.rows and sheet.columns

# =============================================================================
# CODE EXAMPLES
# =============================================================================

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# -----------------------------------------------------------------------------
# CREATING A TEST EXCEL FILE
# -----------------------------------------------------------------------------
# First, we'll create an example Excel file to practice reading from.
# This creates a file called 'example.xlsx' with some sample data.

wb_new = openpyxl.Workbook()
sheet = wb_new.active
sheet['A1'] = 'Name'
sheet['B1'] = 'Age'
sheet['C1'] = 'City'
sheet['A2'] = 'Alice'
sheet['B2'] = 25
sheet['C2'] = 'NYC'
sheet['A3'] = 'Bob'
sheet['B3'] = 30
sheet['C3'] = 'LA'
sheet['A4'] = 'Charlie'
sheet['B4'] = 35
sheet['C4'] = 'Chicago'
wb_new.save('example.xlsx')
print('Created example.xlsx with sample data\n')

# -----------------------------------------------------------------------------
# 1. LOADING WORKBOOKS
# -----------------------------------------------------------------------------
# Use openpyxl.load_workbook() to open an existing Excel file.
# This returns a Workbook object that you can then work with.

wb = openpyxl.load_workbook('example.xlsx')
print(f'Loaded workbook type: {type(wb)}')  
# Output: <class 'openpyxl.workbook.workbook.Workbook'>

# -----------------------------------------------------------------------------
# 2. ACCESSING SHEETS - Getting Sheet Names
# -----------------------------------------------------------------------------
# Every Excel file has one or more worksheets (tabs at the bottom).
# wb.sheetnames gives you a list of all sheet names in the workbook.

print(f'\nSheet names in workbook: {wb.sheetnames}')  
# Output: ['Sheet']

# -----------------------------------------------------------------------------
# 2. ACCESSING SHEETS - Getting a Specific Sheet
# -----------------------------------------------------------------------------
# Use bracket notation wb['SheetName'] to get a specific worksheet.

sheet = wb['Sheet']
print(f'Got sheet: {sheet.title}')  
# Output: Got sheet: Sheet

# -----------------------------------------------------------------------------
# 2. ACCESSING SHEETS - Getting the Active Sheet
# -----------------------------------------------------------------------------
# wb.active gives you whichever sheet was visible when the file was saved.
# This is often the first sheet, but not always.

active_sheet = wb.active
print(f'Active sheet: {active_sheet}')  
# Output: <Worksheet "Sheet">

# -----------------------------------------------------------------------------
# 3. READING CELLS - Method 1: Using Cell Coordinates
# -----------------------------------------------------------------------------
# Use Excel-style coordinates like 'A1', 'B5', 'Z99' to access cells.
# This is intuitive if you're used to Excel.

cell_a1 = sheet['A1']
print(f'\nCell A1 value: {cell_a1.value}')  
# Output: Cell A1 value: Name

# -----------------------------------------------------------------------------
# 4. CELL PROPERTIES
# -----------------------------------------------------------------------------
# Cell objects have useful properties: coordinate, row, column, value

c = sheet['B1']
print(f'\nCell coordinate: {c.coordinate}')  # Output: B1
print(f'Cell row: {c.row}')                  # Output: 1
print(f'Cell column: {c.column}')            # Output: 2
print(f'Cell value: {c.value}')              # Output: Age

# -----------------------------------------------------------------------------
# 3. READING CELLS - Method 2: Using cell() Method
# -----------------------------------------------------------------------------
# Use sheet.cell(row=X, column=Y) to access cells with numbers.
# This is better for loops where you're calculating positions.
# Remember: rows and columns start at 1, not 0!

cell_b2 = sheet.cell(row=2, column=2)
print(f'\nCell B2 value using cell() method: {cell_b2.value}')  
# Output: Cell B2 value using cell() method: 25

# -----------------------------------------------------------------------------
# 5. SHEET DIMENSIONS
# -----------------------------------------------------------------------------
# sheet.max_row tells you the highest row number with data
# sheet.max_column tells you the highest column number with data
# Useful for knowing how much data you have without counting manually

print(f'\nMaximum row with data: {sheet.max_row}')        # Output: 4
print(f'Maximum column with data: {sheet.max_column}')    # Output: 3

# -----------------------------------------------------------------------------
# 6. CONVERTING BETWEEN COLUMN LETTERS AND NUMBERS
# -----------------------------------------------------------------------------
# Excel uses letters (A, B, C...) but programming often needs numbers.
# get_column_letter() converts numbers to letters: 1→A, 2→B, 27→AA
# column_index_from_string() converts letters to numbers: A→1, B→2, AA→27

print(f'\nColumn 1 as letter: {get_column_letter(1)}')        # Output: A
print(f'Column 27 as letter: {get_column_letter(27)}')       # Output: AA
print(f"Column 'A' as number: {column_index_from_string('A')}")   # Output: 1
print(f"Column 'M' as number: {column_index_from_string('M')}")   # Output: 13

# -----------------------------------------------------------------------------
# 7. SLICING - Getting Rectangular Areas of Cells
# -----------------------------------------------------------------------------
# Use sheet['A1':'C3'] to get a rectangular area of cells all at once.
# Returns a tuple of tuples: outer tuple = rows, inner tuples = cells in that row

print('\nSlicing cells A1:C2:')
for row_of_cells in sheet['A1':'C2']:
    for cell in row_of_cells:
        print(f'  {cell.coordinate}: {cell.value}')
    print('  ---')
# Output shows each cell's coordinate and value in that rectangle

# -----------------------------------------------------------------------------
# 8. ITERATING - Looping Through Rows
# -----------------------------------------------------------------------------
# sheet.rows gives you an iterator that produces each row as a tuple of cells.
# Each row is a tuple containing all the Cell objects in that row.

print('\nAll rows (first 3 columns):')
for row in sheet.rows:
    print(f'  {row[0].value}, {row[1].value}, {row[2].value}')
# Output: Name, Age, City
#         Alice, 25, NYC
#         Bob, 30, LA
#         Charlie, 35, Chicago

# -----------------------------------------------------------------------------
# 8. ITERATING - Looping Through Columns
# -----------------------------------------------------------------------------
# sheet.columns gives you an iterator that produces each column as a tuple of cells.
# Each column is a tuple containing all the Cell objects in that column.

print('\nFirst column (all cells):')
first_column = list(sheet.columns)[0]
for cell in first_column:
    print(f'  {cell.value}')
# Output: Name
#         Alice
#         Bob
#         Charlie
