# SECTION 2: WRITING EXCEL FILES - Practice

# Q1: Import openpyxl module
# WHAT IT DOES: Makes openpyxl functions available for working with Excel files
# ┌─ EXAMPLE ─────────────
# │ import sys
# └───────────────────────
import openpyxl




# Q2: Use openpyxl.Workbook() to create a new blank workbook and store it in variable wb
# WHAT IT DOES: Creates a new empty Excel file in memory with one default sheet
# ┌─ EXAMPLE ─────────────
# │ my_wb = openpyxl.Workbook()
# └───────────────────────

wb = openpyxl.Workbook()


# Q3: Use wb.active to get the active sheet and store it in variable sheet
# WHAT IT DOES: Gets the currently active worksheet from the workbook
# ┌─ EXAMPLE ─────────────
# │ my_sheet = my_wb.active
# └───────────────────────
sheet = wb.active



# Q4: Use print() to print sheet.title
# WHAT IT DOES: Displays the name of the worksheet
# ┌─ EXAMPLE ─────────────
# │ print(some_sheet.title)
# └───────────────────────
print(sheet.title)



# Q5: Set sheet.title equal to 'Products'
# WHAT IT DOES: Changes the name of the worksheet
# ┌─ EXAMPLE ─────────────
# │ my_sheet.title = 'Inventory'
# └───────────────────────
sheet.title = 'Products'



# Q6: Use sheet['A1'] to write 'Product' to cell A1 and use sheet['B1'] to write 'Price' to cell B1
# WHAT IT DOES: Sets the value of cells using bracket notation with cell coordinates
# ┌─ EXAMPLE ─────────────
# │ my_sheet['C1'] = 'Quantity'
# └───────────────────────
sheet['A1'] = 'Product'
sheet['B1'] = 'Price'



# Q7: Use sheet['A2'] to write 'Laptop' in cell A2 and use sheet['B2'] to write 999 in cell B2
# WHAT IT DOES: Writes text and number values to specific cells
# ┌─ EXAMPLE ─────────────
# │ my_sheet['C2'] = 50
# └───────────────────────
sheet['A2'] = 'Laptop'
sheet['B2'] = 999


# Q8: Use sheet.cell(row=3, column=1).value to write 'Mouse' in A3, and use sheet.cell(row=3, column=2).value to write 25 in B3
# WHAT IT DOES: Sets cell values using row and column numbers instead of coordinates like 'A3'
# ┌─ EXAMPLE ─────────────
# │ my_sheet.cell(row=4, column=2).value = 15
# └───────────────────────

sheet.cell(row=3, column=1).value = 'Mouse'
sheet.cell(row=3, column=2).value = 25

# Q9: Use wb.create_sheet() with index=0 and title='Summary' to create a new sheet named 'Summary' at the beginning
# WHAT IT DOES: Adds a new worksheet to the workbook at a specific position with a specific name
# ┌─ EXAMPLE ─────────────
# │ my_wb.create_sheet(index=1, title='Details')
# └───────────────────────
wb.create_sheet(index=0, title='Summary')



# Q10: Use print() to print wb.sheetnames
# WHAT IT DOES: Displays a list of all sheet names in the workbook
# ┌─ EXAMPLE ─────────────
# │ print(my_wb.sheetnames)
# └───────────────────────
print(wb.sheetnames)



# Q11: Use wb.save() to save the workbook as 'my_products.xlsx'
# WHAT IT DOES: Writes the workbook from memory to a physical Excel file on disk
# ┌─ EXAMPLE ─────────────
# │ my_wb.save('inventory.xlsx')
# └───────────────────────

wb.save('my_products.xlsx')






