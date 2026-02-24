# SECTION 1: READING EXCEL FILES - Practice

# Q1: Use import to import openpyxl module, and use from openpyxl.utils import to import get_column_letter and column_index_from_string
# WHAT IT DOES: Makes openpyxl functions available plus utilities for converting between column letters and numbers
# ┌─ EXAMPLE ─────────────
# │ import sys
# │ from os import path
# └───────────────────────
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string



# Q2: Use openpyxl.load_workbook() to load 'example.xlsx' and store it in a variable called wb
# WHAT IT DOES: Opens the Excel file and creates a Workbook object you can work with
# ┌─ EXAMPLE ─────────────
# │ workbook = openpyxl.load_workbook('sales.xlsx')
# └───────────────────────
wb = openpyxl.load_workbook('example.xlsx')



# Q3: Use print() to print wb.sheetnames
# WHAT IT DOES: Displays a list of all sheet names in the Excel file
# ┌─ EXAMPLE ─────────────
# │ print(my_workbook.sheetnames)
# └───────────────────────
print(wb.sheetnames)



# Q4: Use wb['Sheet'] to get the sheet named 'Sheet' and store it in a variable called sheet
# WHAT IT DOES: Accesses a specific worksheet by name using bracket notation
# ┌─ EXAMPLE ─────────────
# │ my_sheet = workbook['Sales']
# └───────────────────────
sheet = wb['Sheet']



# Q5: Use print() to print sheet.title
# WHAT IT DOES: Displays the name of the worksheet
# ┌─ EXAMPLE ─────────────
# │ print(my_sheet.title)
# └───────────────────────
print(sheet.title)



# Q6: Use wb.active to get the active sheet and store it in a variable called active_sheet
# WHAT IT DOES: Gets whichever sheet was visible when the file was last saved
# ┌─ EXAMPLE ─────────────
# │ current_sheet = workbook.active
# └───────────────────────
active_sheet = wb.active



# Q7: Use sheet['A1'] to get cell A1 and use print() to print its value using .value
# WHAT IT DOES: Accesses a cell by its coordinate and reads its contents
# ┌─ EXAMPLE ─────────────
# │ cell = my_sheet['B5']
# │ print(cell.value)
# └───────────────────────
cell = sheet['A1']
print(cell.value)



# Q8: Use sheet['B1'] to get cell B1, store it in variable c, then use print() to print c.coordinate, c.row, and c.column
# WHAT IT DOES: Shows how to access cell properties like coordinate, row number, and column number
# ┌─ EXAMPLE ─────────────
# │ my_c = my_sheet['C3']
# │ print(my_c.coordinate, my_c.row, my_c.column)
# └───────────────────────
c = sheet['B1']
print(c.coordinate, c.row, c.column)



# Q9: Use sheet.cell(row=2, column=2) to get cell B2 and use print() to print its value
# WHAT IT DOES: Accesses a cell using row and column numbers instead of coordinate strings
# ┌─ EXAMPLE ─────────────
# │ cell_data = my_sheet.cell(row=3, column=4).value
# │ print(cell_data)
# └───────────────────────
cell_data = sheet.cell(row=2, column=2).value
print(cell_data)



# Q10: Use print() to print sheet.max_row and sheet.max_column
# WHAT IT DOES: Shows how many rows and columns contain data in the sheet
# ┌─ EXAMPLE ─────────────
# │ print('Rows:', my_sheet.max_row)
# │ print('Columns:', my_sheet.max_column)
# └───────────────────────
print('Rows:', sheet.max_row)
print('Columns:', sheet.max_column)



# Q11: Use get_column_letter(14) to convert column number 14 to its letter and use print() to print the result
# WHAT IT DOES: Converts a column number (1, 2, 3...) to Excel's letter format (A, B, C...)
# ┌─ EXAMPLE ─────────────
# │ letter = get_column_letter(1)
# │ print(letter)
# └───────────────────────
letter = get_column_letter(14)
print(letter)



# Q12: Use column_index_from_string('M') to convert column letter 'M' to its number and use print() to print the result
# WHAT IT DOES: Converts Excel's column letter (A, B, C...) to a number (1, 2, 3...)
# ┌─ EXAMPLE ─────────────
# │ number = column_index_from_string('Z')
# │ print(number)
# └───────────────────────
number = column_index_from_string('M')
print(number)



# Q13: Use sheet['A1':'C2'] to get a rectangular area of cells, then use a nested for loop to print each cell's coordinate and value
# WHAT IT DOES: Gets multiple cells at once from a rectangular area like A1 to C2, returns nested tuples
# ┌─ EXAMPLE ─────────────
# │ for row_of_cells in my_sheet['B2':'D3']:
# │     for cell in row_of_cells:
# │         print(cell.coordinate, cell.value)
# └───────────────────────

for row_of_cells in sheet['A1':'C2']:
    for cell in row_of_cells:
        print(cell.coordinate, cell.value)



# Q14: Use sheet.rows with a for loop to iterate through all rows and use print() to print the value of the first cell in each row (row[0].value)
# WHAT IT DOES: Iterates through every row in the sheet, each row is a tuple of cells
# ┌─ EXAMPLE ─────────────
# │ for row in my_sheet.rows:
# │     print(row[0].value)
# └───────────────────────
for row in sheet.rows:
    print(row[0].value)



# Q15: Use sheet.columns with a nested for loop to iterate through all columns and use print() to print each cell's value
# WHAT IT DOES: Iterates through every column in the sheet, each column is a tuple of cells
# ┌─ EXAMPLE ─────────────
# │ for column in my_sheet.columns:
# │     for cell in column:
# │         print(cell.value)
# └───────────────────────

for column in sheet.columns:
    for cell in column:
        print(cell.value)


