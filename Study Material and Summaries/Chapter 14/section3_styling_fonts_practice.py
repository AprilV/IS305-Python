# SECTION 3: STYLING AND FONTS - Practice

# Q1: Write two import statements: First line import openpyxl, second line from openpyxl.styles import Font
# WHAT IT DOES: openpyxl module works with Excel files, Font from openpyxl.styles is used to style cell fonts
# ┌─ EXAMPLE ─────────────
# │ import sys
# │ from os import path
# └───────────────────────
import openpyxl
from openpyxl.styles import Font


# Q2: Use openpyxl.Workbook() to create a new workbook and use .active to get the active sheet, store sheet in variable sheet
# WHAT IT DOES: Workbook() creates a new blank Excel file in memory, .active gets the currently active sheet
# ┌─ EXAMPLE ─────────────
# │ my_wb = openpyxl.Workbook()
# │ my_sheet = my_wb.active
# └───────────────────────
wb = openpyxl.Workbook()
sheet = wb.active



# Q3: Use Font() with size=20 and bold=True to create a Font object and store it in variable header_font
# WHAT IT DOES: Font() creates a font style object that defines how text looks (size, bold, italic, font name)
# ┌─ EXAMPLE ─────────────
# │ title_font = Font(size=16, italic=True)
# └───────────────────────
header_font = Font(size=20, bold=True)



# Q4: Set sheet['A1'].font equal to header_font and use sheet['A1'] to write 'Sales Report' in A1
# WHAT IT DOES: Assigning to .font applies the font style to a cell, then assigning text sets the cell's value
# ┌─ EXAMPLE ─────────────
# │ my_sheet['B1'].font = title_font
# │ my_sheet['B1'] = 'Monthly Data'
# └───────────────────────
sheet['A1'].font = header_font
sheet['A1'] = 'Sales report'


# Q5: Use Font() with name='Times New Roman' and italic=True to create a Font object and store it in variable body_font
# WHAT IT DOES: Font() with name parameter changes the font family (like Times New Roman, Arial, Calibri)
# ┌─ EXAMPLE ─────────────
# │ data_font = Font(name='Arial', bold=True)
# └───────────────────────
body_font = Font(name='Times New Roman', italic=True)



# Q6: Set sheet['A3'].font equal to body_font and use sheet['A3'] to write 'Report generated on 2/23/2026' in A3
# WHAT IT DOES: Applies the body_font style to cell A3 and then sets the text content of that cell
# ┌─ EXAMPLE ─────────────
# │ my_sheet['B3'].font = data_font
# │ my_sheet['B3'] = 'Data entry'
# └───────────────────────
sheet['A3'].font = body_font
sheet['A3'] = 'Report Generated on 2/23/2026'



# Q7: Set sheet.row_dimensions[1].height equal to 40
# WHAT IT DOES: row_dimensions[1] gets row 1, and .height sets the row height in points (1/72 of an inch)
# ┌─ EXAMPLE ─────────────
# │ my_sheet.row_dimensions[2].height = 25
# └───────────────────────
sheet.row_dimensions[1].height = 40



# Q8: Set sheet.column_dimensions['A'].width equal to 30
# WHAT IT DOES: column_dimensions['A'] gets column A, and .width sets how many characters wide the column is
# ┌─ EXAMPLE ─────────────
# │ my_sheet.column_dimensions['B'].width = 20
# └───────────────────────
sheet.column_dimensions['A'].width = 30



# Q9: Use wb.save() to save the workbook as 'styled_report.xlsx'
# WHAT IT DOES: .save() writes the workbook in memory to a physical .xlsx file on disk
# ┌─ EXAMPLE ─────────────
# │ my_wb.save('formatted.xlsx')
# └───────────────────────

wb.save('styled_report.xlsx')


