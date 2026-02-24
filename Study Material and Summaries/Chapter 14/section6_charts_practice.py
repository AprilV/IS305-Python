# SECTION 6: CHARTS - Practice

# Q1: Import openpyxl module and use from openpyxl.chart import to import BarChart, Reference, and Series
# WHAT IT DOES: Makes openpyxl available plus chart objects needed to create visual charts in Excel
# ┌─ EXAMPLE ─────────────
# │ import sys
# │ from os import path
# └───────────────────────
import openpyxl
from openpyxl.chart import BarChart, Reference, Series



# Q2: Use openpyxl.Workbook() to create a new workbook and use .active to get the active sheet, store sheet in variable sheet
# WHAT IT DOES: Creates a new empty Excel file in memory and gets its active worksheet
# ┌─ EXAMPLE ─────────────
# │ my_wb = openpyxl.Workbook()
# │ my_sheet = my_wb.active
# └───────────────────────
wb = openpyxl.Workbook()
sheet = wb.active



# Q3: Use sheet['A1'] to write 'Product' and use sheet['B1'] to write 'Units Sold'
# WHAT IT DOES: Writes column headers for the data that will be charted
# ┌─ EXAMPLE ─────────────
# │ my_sheet['A1'] = 'Month'
# │ my_sheet['B1'] = 'Revenue'
# └───────────────────────

sheet['A1'] = 'Product'
sheet['B1'] = 'Revenue'


# Q4: Use sheet['A2'] to write 'Laptops' and use sheet['B2'] to write 45
# WHAT IT DOES: Writes the first row of data (product name and quantity) that will appear in the chart
# ┌─ EXAMPLE ─────────────
# │ my_sheet['A2'] = 'Tablets'
# │ my_sheet['B2'] = 30
# └───────────────────────
sheet['A2'] = 'Laptops'
sheet['B2'] = 45



# Q5: Use sheet['A3'] to write 'Phones' and use sheet['B3'] to write 78, then use sheet['A4'] to write 'Tablets' and use sheet['B4'] to write 62
# WHAT IT DOES: Writes two more rows of data (more products and quantities) for the chart
# ┌─ EXAMPLE ─────────────
# │ my_sheet['A3'] = 'Monitors'
# │ my_sheet['B3'] = 15
# └───────────────────────
sheet['A3'] = 'Phones'
sheet['B3'] = 78
sheet['A4'] = 'Tablets'
sheet['B4'] = 62

# Q6: Create a Reference object that points to cells B2, B3, B4 (the three numbers you wrote: 45, 78, 62)
# Use Reference(sheet, min_row=2, max_row=4, min_col=2, max_col=2) and store in variable refObj
# WHAT IT DOES: Reference tells Python which cells contain the numbers for your chart bars
# min_row=2 means start at row 2, max_row=4 means end at row 4
# min_col=2 and max_col=2 means use column B only (A=1, B=2, C=3, etc.)
# So this grabs B2:B4 which contains your three numbers
# ┌─ EXAMPLE ─────────────
# │ my_ref = Reference(my_sheet, min_row=1, max_row=3, min_col=1, max_col=1)
# └───────────────────────
refobj = Reference(sheet, min_row=2, max_row=4, min_col=2, max_col=2)



# Q7: Use Series() with refObj and title='Sales Data' to create a Series object and store in variable seriesObj
# WHAT IT DOES: Creates a data series from the reference that will be displayed in the chart with a label
# ┌─ EXAMPLE ─────────────
# │ my_series = Series(my_ref, title='Monthly Totals')
# └───────────────────────
seriesObj = Series(refobj, title='Sales Data')



# Q8: Use BarChart() to create a BarChart object and store in variable chartObj
# WHAT IT DOES: Creates a bar chart object that will display data as vertical or horizontal bars
# ┌─ EXAMPLE ─────────────
# │ my_chart = BarChart()
# └───────────────────────
chartObj = BarChart()



# Q9: Set chartObj.title equal to 'Product Sales Chart'
# WHAT IT DOES: Sets the title text that will appear at the top of the chart
# ┌─ EXAMPLE ─────────────
# │ my_chart.title = 'Revenue Overview'
# └───────────────────────
chartObj.title = 'Product Sales Chart'



# Q10: Use chartObj.append() to add seriesObj to the chart
# WHAT IT DOES: Adds the data series to the chart so the chart knows what data to display
# ┌─ EXAMPLE ─────────────
# │ my_chart.append(my_series)
# └───────────────────────
chartObj.append(seriesObj)



# Q11: Use sheet.add_chart() with chartObj and 'D2' to add the chart to the sheet at cell D2
# WHAT IT DOES: Places the chart onto the worksheet with its top-left corner at the specified cell
# ┌─ EXAMPLE ─────────────
# │ my_sheet.add_chart(my_chart, 'E5')
# └───────────────────────
sheet.add_chart(chartObj, 'D2')



# Q12: Use wb.save() to save the workbook as 'chart_example.xlsx'
# WHAT IT DOES: Writes the workbook with the embedded chart to a physical Excel file on disk
# ┌─ EXAMPLE ─────────────
# │ my_wb.save('charted_data.xlsx')
# └───────────────────────
wb.save('chart_examples.xlsx')



